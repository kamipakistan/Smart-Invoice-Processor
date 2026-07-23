import io
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any

def generate_flat_excel_report(invoices_data: List[Dict[str, Any]]) -> bytes:
    """
    Consolidates list of invoice headers and nested line items into a single 16-column flat Excel spreadsheet.
    Fields 1-5 (Header) are duplicated for each line item (Fields 6-16).
    """
    flat_rows = []
    
    for inv in invoices_data:
        hdr = {
            "FBR Invoice No": inv.get("fbr_invoice_no") or "",
            "Registration No": inv.get("registration_no") or "",
            "Business Name": inv.get("business_name") or "",
            "Invoice Date": inv.get("invoice_date") or "",
            "Insertion Date": inv.get("insertion_date") or ""
        }
        
        line_items = inv.get("line_items", [])
        if not line_items:
            row = {
                "Sr. No": "",
                "FBR Invoice No": hdr["FBR Invoice No"],
                "Registration No": hdr["Registration No"],
                "Business Name": hdr["Business Name"],
                "Invoice Date": hdr["Invoice Date"],
                "Insertion Date": hdr["Insertion Date"],
                "HS Code": "",
                "Product Description": "",
                "Sales Type": "",
                "Quantity": 0,
                "UoM": "",
                "Sales Value": 0.0,
                "Retail Price": 0.0,
                "Sales Tax": 0.0,
                "Further Tax": 0.0,
                "FED": 0.0
            }
            flat_rows.append(row)
        else:
            for item in line_items:
                row = {
                    "Sr. No": item.get("sr_no") if item.get("sr_no") is not None else "",
                    "FBR Invoice No": hdr["FBR Invoice No"],
                    "Registration No": hdr["Registration No"],
                    "Business Name": hdr["Business Name"],
                    "Invoice Date": hdr["Invoice Date"],
                    "Insertion Date": hdr["Insertion Date"],
                    "HS Code": item.get("hs_code") or "",
                    "Product Description": item.get("product_description") or "",
                    "Sales Type": item.get("sales_type") or "",
                    "Quantity": float(item.get("quantity") or 0.0),
                    "UoM": item.get("uom") or "",
                    "Sales Value": float(item.get("sales_value") or 0.0),
                    "Retail Price": float(item.get("retail_price") or 0.0),
                    "Sales Tax": float(item.get("sales_tax") or 0.0),
                    "Further Tax": float(item.get("further_tax") or 0.0),
                    "FED": float(item.get("fed") or 0.0)
                }
                flat_rows.append(row)

    df = pd.DataFrame(flat_rows)

    columns_order = [
        "Sr. No",
        "FBR Invoice No",
        "Registration No",
        "Business Name",
        "Invoice Date",
        "Insertion Date",
        "HS Code",
        "Product Description",
        "Sales Type",
        "Quantity",
        "UoM",
        "Sales Value",
        "Retail Price",
        "Sales Tax",
        "Further Tax",
        "FED"
    ]
    
    for col in columns_order:
        if col not in df.columns:
            df[col] = ""
            
    df = df[columns_order]

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="FBR Invoices")
        workbook = writer.book
        worksheet = writer.sheets["FBR Invoices"]

        # Professional OpenPyXL styling
        header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid") # Dark slate
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        regular_font = Font(name="Calibri", size=10)
        thin_border = Border(
            left=Side(style='thin', color='E5E7EB'),
            right=Side(style='thin', color='E5E7EB'),
            top=Side(style='thin', color='E5E7EB'),
            bottom=Side(style='thin', color='E5E7EB')
        )

        for col_idx, col in enumerate(columns_order, 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in worksheet.iter_rows(min_row=2, max_row=len(flat_rows) + 1, min_col=1, max_col=16):
            for cell in row:
                cell.font = regular_font
                cell.border = thin_border
                if cell.column_letter in ['J', 'L', 'M', 'N', 'O', 'P']: # Numeric columns
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        # Auto-adjust column widths
        for col in worksheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            worksheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

    return output.getvalue()
